import io
from datetime import datetime
from fastapi import APIRouter, Depends, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from ..deps import get_current_user
from ..database import get_db
from ..models import Transaction, Debt, Account, Category
from ..response import ok

router = APIRouter(prefix="/data", tags=["data"])

@router.get("/export")
def export_data(db: Session = Depends(get_db), user=Depends(get_current_user)):
    wb = Workbook()
    ws_trx = wb.active
    ws_trx.title = "transactions"
    ws_trx.append(["transaction_id", "type", "amount", "category", "account", "time", "summary", "person"])
    rows = db.query(Transaction, Category.category_name, Account.account_name).join(Category, Transaction.category_id == Category.category_id).join(Account, Transaction.account_id == Account.account_id).filter(Transaction.user_id == user.user_id).order_by(Transaction.transaction_time).all()
    for r in rows:
        ws_trx.append([
            r.Transaction.transaction_id,
            r.Transaction.transaction_type,
            float(r.Transaction.amount),
            r[1],
            r[2],
            r.Transaction.transaction_time.strftime("%Y-%m-%d %H:%M:%S"),
            r.Transaction.summary or "",
            r.Transaction.target_person or "",
        ])

    ws_debts = wb.create_sheet("debts")
    ws_debts.append(["debt_id", "type", "person", "amount", "time", "note"])
    drows = db.query(Debt).filter(Debt.user_id == user.user_id).order_by(Debt.action_time).all()
    for d in drows:
        ws_debts.append([
            d.debt_id,
            d.debt_type,
            d.person_name,
            float(d.amount),
            d.action_time.strftime("%Y-%m-%d %H:%M:%S"),
            d.note or "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"meloonmoney_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@router.post("/import")
def import_data(file: UploadFile, db: Session = Depends(get_db), user=Depends(get_current_user)):
    content = file.file.read()
    wb = load_workbook(io.BytesIO(content))
    if "transactions" in wb.sheetnames:
        ws = wb["transactions"]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            _, ttype, amount, category_name, account_name, ttime, summary, person = row
            cat = db.query(Category).filter(Category.user_id == user.user_id, Category.category_name == category_name).first()
            acc = db.query(Account).filter(Account.user_id == user.user_id, Account.account_name == account_name).first()
            if not cat or not acc:
                continue
            dt = datetime.strptime(ttime, "%Y-%m-%d %H:%M:%S") if isinstance(ttime, str) else datetime.fromtimestamp(ttime.timestamp())
            trx = Transaction(user_id=user.user_id, account_id=acc.account_id, category_id=cat.category_id, amount=float(amount), transaction_type=ttype, transaction_time=dt, summary=summary, target_person=person)
            db.add(trx)
        db.commit()
    if "debts" in wb.sheetnames:
        ws = wb["debts"]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            _, dtype, person, amount, dtime, note = row
            dt = datetime.strptime(dtime, "%Y-%m-%d %H:%M:%S") if isinstance(dtime, str) else datetime.fromtimestamp(dtime.timestamp())
            d = Debt(user_id=user.user_id, debt_type=dtype, person_name=person, amount=float(amount), action_time=dt, note=note)
            db.add(d)
        db.commit()
    return ok(message="导入完成")
